from datetime import datetime, timezone
from typing import Any

from django.contrib.auth.decorators import login_required
from django.core.exceptions import PermissionDenied
from django.http import (
    HttpRequest,
    HttpResponse,
    HttpResponsePermanentRedirect,
    HttpResponseRedirect,
)
from django.shortcuts import redirect, render
from django.urls import reverse_lazy
from django.utils import timezone as d_timezone
from django.utils.dateparse import parse_datetime
from django.views.generic import ListView
from django.views.generic.edit import CreateView, UpdateView

from django_stubs_ext import QuerySetAny

from .forms import (
    CreateMalfunctionsForm,
    UpdateModelMalfunctionsForm,
    FilterForm,
)
from .models import ModelMalfunctions

from openpyxl import load_workbook

from io import BytesIO


class CalculationOfDowntime:
    def __init__(self, filter_model: ModelMalfunctions) -> None:
        self.text_hours: str = str()
        self.text_minutes: str = str()
        self.model = filter_model
        self.dt_closed: datetime = (
            filter_model.date_time_closed
            if filter_model.date_time_closed is not None
            else datetime.now()
        )
        self.dt_accepted: datetime = (
            filter_model.date_time_accepted
            if filter_model.date_time_closed is not None
            else datetime.now()
        )
        return

    def cod(self) -> str:
        if self.dt_closed > self.dt_accepted:
            downtime_sec = (self.dt_closed - self.dt_accepted).total_seconds()
            downtime_min = int(downtime_sec / 60)
            full_hours = int(downtime_min // 60)
            minutes = int(downtime_min - (full_hours * 60))
            if full_hours != 0:
                self.text_hours = f"{full_hours} ч."
            if minutes != 0:
                self.text_minutes = f"{minutes} мин."
            return f"{self.text_hours} {self.text_minutes}"
        else:
            return "ОШИБКА! Время закрытия заявки, раньше её создания."


class MalfunctionsList(ListView):
    model: Any = ModelMalfunctions.objects
    template_name = "malfunctions/list.html"
    context_object_name = "list"

    def get_queryset(
        self,
    ) -> QuerySetAny[ModelMalfunctions, ModelMalfunctions]:
        return ModelMalfunctions.objects.all().order_by("-date_time_accepted")


class MalfunctionsUpdate(UpdateView):
    model = ModelMalfunctions
    form_class = UpdateModelMalfunctionsForm
    template_name = "malfunctions/edit.html"
    pk_url_kwarg = "info_id"
    context_object_name = "info"
    success_url = reverse_lazy("list")

    def form_valid(self, form: UpdateModelMalfunctionsForm) -> HttpResponse:
        report = form.save(commit=False)

        # Рассчитываем время простоя, если есть время завершения
        if report.date_time_closed:
            report.simple = CalculationOfDowntime(filter_model=report).cod()
        report.save()

        return super().form_valid(form)


@login_required
def delete_contact(
    request: HttpRequest, info_id: int
) -> HttpResponseRedirect | HttpResponsePermanentRedirect:
    info = ModelMalfunctions.objects.get(pk=info_id)
    if request.user.is_superuser:
        info.delete()
        return redirect("list")
    raise PermissionDenied()


@login_required
def send_archive(
    request: HttpRequest, info_id: int
) -> HttpResponseRedirect | HttpResponsePermanentRedirect:
    info = ModelMalfunctions.objects.get(pk=info_id)
    if request.user.is_superuser:
        # изменяем значение поля status на False
        info.status = False
        info.date_time_closed = datetime.now().astimezone(timezone.utc)
        # Рассчитываем время простоя, если есть время завершения
        info.simple = CalculationOfDowntime(filter_model=info).cod()
        # сохраняем изменения в базе данных
        info.save()
        return redirect("list")
    raise PermissionDenied()


@login_required
def send_black(
    request: HttpRequest, info_id: int
) -> HttpResponseRedirect | HttpResponsePermanentRedirect:
    info = ModelMalfunctions.objects.get(pk=info_id)
    if request.user.is_superuser:
        # изменяем значение поля status на False
        info.status = True
        # сохраняем изменения в базе данных
        info.save()
        return redirect("list")
    raise PermissionDenied()


class MalfunctionsCreate(CreateView):
    model = ModelMalfunctions
    form_class = CreateMalfunctionsForm
    template_name = "malfunctions/create.html"
    success_url = reverse_lazy("list")

    def form_valid(self, form: CreateMalfunctionsForm) -> HttpResponse:
        report = form.save(commit=False)
        report.save()

        # Рассчитываем время простоя, если есть время завершения
        if report.date_time_closed:
            downtime = report.date_time_closed - report.date_time_accepted
            # downtime_minutes теперь содержит разницу во времени между
            # завершением и началом работ
            # Можете сохранить значение в минутах или в нужном вам формате
            report.simple = int(downtime.total_seconds() / 60)
            report.save()

        return super().form_valid(form)

    def form_invalid(self, form: CreateMalfunctionsForm) -> HttpResponse:
        print("что-то не так!")
        # Добавьте здесь необходимые действия для обработки невалидной формы
        errors = form.errors.as_data()
        print(errors)
        # Теперь переменная errors содержит информацию о том,
        # почему форма невалидна
        return self.render_to_response(
            self.get_context_data(form=form, errors=errors)
        )


@login_required
def filter_malfunctions(request: HttpRequest) -> HttpResponse:
    # Преобразование значений в формат "%Y-%m-%dT%H:%M"
    start_datetime = parse_datetime(request.GET.get("start_datetime", ""))
    end_datetime = parse_datetime(request.GET.get("end_datetime", ""))
    address = request.GET.get("address", "")
    num_house = request.GET.get("num_house", "")
    entrance = request.GET.get("entrance", "")
    mechanics = request.GET.get("mechanics", "")
    status = request.GET.get("status", "")

    print(f"{mechanics=}")
    print(f"{status=}")

    # Создание формы с переданными значениями
    form_class = FilterForm(
        {
            "start_datetime": start_datetime,
            "end_datetime": end_datetime,
            "address": address,
            "num_house": num_house,
            "entrance": entrance,
            "mechanics": mechanics,
            "status": status
        }
    )
    if form_class.is_valid():
        # Получение объектов согласно условиям фильтрации
        filtered_objects = ModelMalfunctions.objects.filter(
            # Больше или равно начальной дате и времени
            date_time_accepted__gte=start_datetime,
            # Меньше или равно конечной дате и времени
            date_time_accepted__lte=end_datetime,
        ).order_by("-date_time_accepted")
        if address != "":
            filtered_objects = filtered_objects.filter(address=address)
        if num_house != "":
            filtered_objects = filtered_objects.filter(num_house=num_house)
        if entrance != "":
            filtered_objects = filtered_objects.filter(entrance=entrance)
        if mechanics != "":
            filtered_objects = filtered_objects.filter(mechanics=mechanics)
        if status not in ["", "all"]:
            filtered_objects = filtered_objects.filter(status=status)
    else:
        filtered_objects = ModelMalfunctions.objects.all().order_by(
            "-date_time_accepted"
        )
        form_class = FilterForm()
    if "download" in request.GET and request.GET["download"] == "true":
        # Загрузка шаблона Excel
        wb = load_workbook(filename="media/data/malfunctions_template.xlsx")
        ws = wb.active

        # Заполнение данных из объектов в шаблон Excel
        row = 3  # начинаем со второй строки, так как первая строка содержит заголовки
        for item in filtered_objects:
            mechanic_fio = "\n".join(
                mechanic.fio for mechanic in item.mechanics.all()
            )
            ws.cell(row=row, column=1, value=row - 2)
            ws.cell(row=row, column=2, value=item.address.address)
            ws.cell(row=row, column=3, value=item.num_house)
            ws.cell(row=row, column=4, value=item.entrance)
            ws.cell(row=row, column=5, value=item.flat_or_tel)
            ws.cell(row=row, column=6, value=item.dispatcher.fio)
            ws.cell(
                row=row,
                column=7,
                value=d_timezone.localtime(item.date_time_accepted).replace(
                    tzinfo=None
                ),
            )
            ws.cell(row=row, column=8, value=mechanic_fio)
            ws.cell(
                row=row,
                column=9,
                value=d_timezone.localtime(item.date_time_transfer).replace(
                    tzinfo=None
                ),
            )
            ws.cell(
                row=row,
                column=10,
                value=d_timezone.localtime(item.date_time_closed).replace(
                    tzinfo=None
                ),
            )
            ws.cell(row=row, column=11, value=item.simple)
            ws.cell(row=row, column=12, value=item.malfunction_and_cause)
            ws.cell(row=row, column=13, value=item.description)
            # Добавьте заполнение других полей объекта, если необходимо
            row += 1

        # Создание HTTP-ответа с содержимым файла Excel
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response[
            "Content-Disposition"
        ] = 'attachment; filename="malfunctions.xlsx"'

        # Сохранение книги Excel в байтовый поток и запись его в HTTP-ответ
        output = BytesIO()
        wb.save(output)
        response.write(output.getvalue())
        return response
    return render(
        request,
        "malfunctions/filter_malfunction.html",
        context={"form": form_class, "list": filtered_objects},
    )


@login_required
def export_excel(request):
    if request.method == "GET":
        start_datetime = parse_datetime(request.GET.get("start_datetime", ""))
        end_datetime = parse_datetime(request.GET.get("end_datetime", ""))

        # Создание формы с переданными значениями
        form_class = FilterForm(
            {"start_datetime": start_datetime, "end_datetime": end_datetime}
        )
        if form_class.is_valid():
            # Получение объектов согласно условиям фильтрации
            filtered_objects = ModelMalfunctions.objects.filter(
                # Больше или равно начальной дате и времени
                date_time_accepted__gte=start_datetime,
                # Меньше или равно конечной дате и времени
                date_time_accepted__lte=end_datetime,
            ).order_by("-date_time_accepted")
        else:
            filtered_objects = ModelMalfunctions.objects.all().order_by(
                "-date_time_accepted"
            )

        # Загрузка шаблона Excel
        wb = load_workbook(filename="media/data/malfunctions_template.xlsx")
        ws = wb.active

        # Заполнение данных из объектов в шаблон Excel
        row = 3  # начинаем со второй строки, так как первая строка содержит заголовки
        for item in filtered_objects:
            mechanic_fio = "\n".join(
                mechanic.fio for mechanic in item.mechanics.all()
            )
            ws.cell(row=row, column=1, value=item.id)
            ws.cell(row=row, column=2, value=item.address.address)
            ws.cell(row=row, column=3, value=item.num_house)
            ws.cell(row=row, column=4, value=item.entrance)
            ws.cell(row=row, column=5, value=item.flat_or_tel)
            ws.cell(row=row, column=6, value=item.dispatcher.fio)
            ws.cell(
                row=row,
                column=7,
                value=d_timezone.localtime(item.date_time_accepted).replace(
                    tzinfo=None
                ),
            )
            ws.cell(row=row, column=8, value=mechanic_fio)
            ws.cell(
                row=row,
                column=9,
                value=d_timezone.localtime(item.date_time_transfer).replace(
                    tzinfo=None
                ),
            )
            ws.cell(
                row=row,
                column=10,
                value=d_timezone.localtime(item.date_time_closed).replace(
                    tzinfo=None
                ),
            )
            ws.cell(row=row, column=11, value=item.simple)
            ws.cell(row=row, column=12, value=item.malfunction_and_cause)
            ws.cell(row=row, column=13, value=item.description)
            # Добавьте заполнение других полей объекта, если необходимо
            row += 1

        # Создание HTTP-ответа с содержимым файла Excel
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response[
            "Content-Disposition"
        ] = 'attachment; filename="malfunctions.xlsx"'

        # Сохранение книги Excel в байтовый поток и запись его в HTTP-ответ
        output = BytesIO()
        wb.save(output)
        response.write(output.getvalue())
        return response
